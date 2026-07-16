
# generate_training_labels.py
import json
import re
import logging
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Any, Tuple, Set

import openpyxl
from tqdm import tqdm
from rapidfuzz import fuzz

# ==============================================================================
# CONFIGURATION
# ==============================================================================

DATA_FILE = Path("dataset/new_resumes.xlsx") 
TEXT_DIR = Path("dataset/new_normalized_text")
OUTPUT_FILE = Path("dataset/new_ner_training_data.jsonl")

# Reporting files
UNMATCHED_REPORT = Path("unmatched_entities.csv")
COVERAGE_REPORT = Path("coverage_report.csv")
LABEL_STATS_REPORT = Path("label_statistics.csv")

# Toggles
ENABLE_SYNONYMS = True
ENABLE_MULTI_MATCH = True
ENABLE_FUZZY = True

MAX_SPAN_LEN = {
    "NAME": 40,
    "EMAIL": 60,
    "PHONE": 20,
    "LOCATION": 40,
    "DEGREE": 50,
    "ORG": 60,
    "DATE": 25,
    "TITLE": 50,
    "CERTIFICATION": 60,
    "SKILL": 35,
}
DEFAULT_MAX_LEN = 60

FUZZY_THRESHOLDS = {
    "NAME": 98,
    "ORG": 95,
    "SKILL": 95,
    "TITLE": 95,
    "CERTIFICATION": 95,
    "DEGREE": 95,
    "LOCATION": 95
}
DEFAULT_FUZZY_THRESHOLD = 95

# ==============================================================================
# LOGGING SETUP
# ==============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)

# ==============================================================================
# DATA STRUCTURES
# ==============================================================================

@dataclass
class EntitySpan:
    start: int
    end: int
    label: str
    text: str
    method: str
    confidence: float

    def to_spacy_format(self) -> List[Any]:
        return [self.start, self.end, self.label]

@dataclass
class MatchStats:
    total_resumes: int = 0
    processed_resumes: int = 0
    skipped_resumes: int = 0
    entities_generated: int = 0
    entities_skipped: int = 0
    exact_matches: int = 0
    regex_matches: int = 0
    synonym_matches: int = 0
    fuzzy_matches: int = 0
    label_counts: Dict[str, int] = None
    unmatched_values: List[Tuple[str, str, str]] = None 

    def __post_init__(self):
        self.label_counts = defaultdict(int)
        self.unmatched_values = []

# ==============================================================================
# NORMALIZATION & VALIDATION
# ==============================================================================

class TextNormalizer:
    @staticmethod
    def normalize_text(text: str) -> str:
        if not text: return ""
        text = unicodedata.normalize("NFKC", text)
        text = re.sub(r'[\u2018\u2019\u00b4\u0060]', "'", text)
        text = re.sub(r'[\u201c\u201d]', '"', text)
        text = text.replace('\xa0', ' ')
        text = re.sub(r'[\u200b\u200e\u200f\uFEFF]', '', text)
        text = re.sub(r'[\u2022\u2023\u25E6\u2043\u2219]', '-', text)
        text = re.sub(r'[ \t]+', ' ', text)
        return text.strip()

    @staticmethod
    def refine_span(text: str, start: int, end: int) -> Tuple[int, int]:
        bad_chars = " \t\n\r,.:;-—|/()[]*"
        while start < end and text[start] in bad_chars:
            start += 1
        while end > start and text[end - 1] in bad_chars:
            end -= 1
        return start, end

class EntityValidator:
    EMAIL_REGEX = re.compile(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$")
    PHONE_REGEX = re.compile(r"^\+?[\d\s\-\(\)]{7,20}$")
    BLACKLIST = ["designation", "email", "contact", "role", "id", "phone", "skills"]

    @classmethod
    def is_valid(cls, value: str, label: str) -> bool:
        if not value or not str(value).strip(): return False
        value = str(value).strip()
        
        max_len = MAX_SPAN_LEN.get(label, DEFAULT_MAX_LEN)
        if len(value) > max_len or len(value) < 2:
            return False

        val_lower = value.lower()
        if any(word in val_lower for word in cls.BLACKLIST) and label not in ["SKILL", "TITLE"]:
             return False
        if "designation" in val_lower or "role" in val_lower:
            return False

        if label == "EMAIL": return bool(cls.EMAIL_REGEX.match(value))
        if label == "PHONE":
            digits = re.sub(r'\D', '', value)
            return 7 <= len(digits) <= 15 and bool(cls.PHONE_REGEX.match(value))
        if label == "DATE":
            if not re.search(r'\d|Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec', value, re.IGNORECASE):
                return False
        
        return True

# ==============================================================================
# MATCHER ENGINE
# ==============================================================================

class Matcher:
    def __init__(self, text: str):
        self.text = text
        self.text_lower = text.lower()
        self.lines = [(m.start(), m.end(), m.group(0)) for m in re.finditer(r'[^\n]+', self.text)]

    def _find_via_regex(self, pattern: str, label: str, method: str, confidence: float) -> List[EntitySpan]:
        spans = []
        try:
            for match in re.finditer(pattern, self.text, re.IGNORECASE):
                start, end = TextNormalizer.refine_span(self.text, match.start(), match.end())
                if start >= end: continue
                
                matched_str = self.text[start:end]
                if EntityValidator.is_valid(matched_str, label):
                    spans.append(EntitySpan(start, end, label, matched_str, method, confidence))
                if not ENABLE_MULTI_MATCH: break
        except re.error: pass
        return spans

    def exact_match(self, value: str, label: str) -> List[EntitySpan]:
        pattern = r'\b' + re.escape(str(value)) + r'\b'
        if not re.search(pattern, self.text, re.IGNORECASE):
            pattern = re.escape(str(value))
        return self._find_via_regex(pattern, label, "EXACT", 100.0)

    def regex_match(self, value: str, label: str) -> List[EntitySpan]:
        words = [re.escape(w) for w in str(value).split()]
        if not words: return []
        pattern = r'[\s\-\.,]*'.join(words)
        return self._find_via_regex(pattern, label, "REGEX", 95.0)

    def fuzzy_match(self, value: str, label: str) -> List[EntitySpan]:
        if not ENABLE_FUZZY: return []
        threshold = FUZZY_THRESHOLDS.get(label, DEFAULT_FUZZY_THRESHOLD)
        best_spans = []
        
        for start_idx, end_idx, line in self.lines:
            if not line.strip(): continue
            score = fuzz.partial_ratio(str(value).lower(), line.lower())
            
            if score >= threshold:
                tokens = line.split()
                val_tokens = str(value).split()
                target_len = len(val_tokens)
                best_sub_score, best_sub_text = 0, ""
                
                for window_size in [target_len - 1, target_len, target_len + 1]:
                    if window_size <= 0: continue
                    for i in range(len(tokens) - window_size + 1):
                        sub_text = " ".join(tokens[i:i+window_size])
                        sub_score = fuzz.ratio(str(value).lower(), sub_text.lower())
                        if sub_score > best_sub_score:
                            best_sub_score, best_sub_text = sub_score, sub_text
                
                if best_sub_text and best_sub_score >= threshold - 5: 
                    sub_match = re.search(re.escape(best_sub_text), line, re.IGNORECASE)
                    if sub_match:
                        abs_start, abs_end = TextNormalizer.refine_span(
                            self.text, start_idx + sub_match.start(), start_idx + sub_match.end()
                        )
                        if abs_start < abs_end and EntityValidator.is_valid(self.text[abs_start:abs_end], label):
                            best_spans.append(EntitySpan(
                                abs_start, abs_end, label, self.text[abs_start:abs_end], "FUZZY", score
                            ))
        
        best_spans.sort(key=lambda x: x.confidence, reverse=True)
        if not ENABLE_MULTI_MATCH and best_spans: return [best_spans[0]]
        return best_spans

    def find_all(self, value: str, label: str) -> List[EntitySpan]:
        if value is None or not str(value).strip(): return []
        value = str(value).strip()
        if not EntityValidator.is_valid(value, label): return []

        spans = self.exact_match(value, label)
        if not spans: spans = self.regex_match(value, label)
        if not spans: spans = self.fuzzy_match(value, label)
        return spans

# ==============================================================================
# RESOLVER
# ==============================================================================

def resolve_overlaps(entities: List[EntitySpan]) -> List[EntitySpan]:
    LABEL_PRIORITY = {"NAME": 10, "EMAIL": 9, "PHONE": 9, "DATE": 8, "CERTIFICATION": 7, "DEGREE": 7, "ORG": 6, "TITLE": 5, "LOCATION": 4, "SKILL": 3}
    METHOD_PRIORITY = {"EXACT": 4, "SYNONYM": 3, "REGEX": 2, "FUZZY": 1}

    sorted_entities = sorted(entities, key=lambda e: (
        e.confidence, METHOD_PRIORITY.get(e.method, 0), (e.end - e.start), LABEL_PRIORITY.get(e.label, 0)
    ), reverse=True)

    resolved, used_indices = [], set()
    for entity in sorted_entities:
        entity_range = set(range(entity.start, entity.end))
        if not entity_range.intersection(used_indices):
            resolved.append(entity)
            used_indices.update(entity_range)

    return sorted(resolved, key=lambda e: e.start)

# ==============================================================================
# PARSING & ORCHESTRATION
# ==============================================================================

def process_dataset():
    if not DATA_FILE.exists():
        logger.error(f"File not found: {DATA_FILE}")
        return
        
    logger.info(f"Loading Excel data from {DATA_FILE}...")
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb.active  # Get the active/first sheet
    
    # Read headers
    headers = [cell.value for cell in ws[1]]
    
    # Read data rows into a list of dicts
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip completely empty rows
            data.append(dict(zip(headers, row)))

    stats, examples = MatchStats(total_resumes=len(data)), []
    
    for row in tqdm(data, desc="Processing Resumes"):
        file_name = row.get("File Name") or row.get("file_name") or row.get("Filename")
        if not file_name: continue

        text_file = TEXT_DIR / f"{Path(file_name).stem}.txt"
        if not text_file.exists():
            stats.skipped_resumes += 1
            continue

        matcher = Matcher(TextNormalizer.normalize_text(text_file.read_text(encoding="utf-8")))
        all_spans: List[EntitySpan] = []

        def _extract(val: Any, label: str):
            if val and (spans := matcher.find_all(str(val), label)):
                all_spans.extend(spans)
            elif val:
                stats.entities_skipped += 1
                stats.unmatched_values.append((file_name, label, str(val)))

        _extract(row.get("Name"), "NAME")
        _extract(row.get("Email"), "EMAIL")
        _extract(row.get("Phone"), "PHONE")
        _extract(row.get("Location"), "LOCATION")
        
        skills_raw = row.get("Skills")
        if skills_raw:
            for skill in str(skills_raw).split(";"): _extract(skill.strip(), "SKILL")

        _extract(row.get("Degree"), "DEGREE")
        _extract(row.get("Institution"), "ORG")
        _extract(row.get("Start Year"), "DATE")
        _extract(row.get("End Year"), "DATE")
        _extract(row.get("Company"), "ORG")
        _extract(row.get("Designation"), "TITLE")
        _extract(row.get("Start Date"), "DATE")
        _extract(row.get("End Date"), "DATE")
        _extract(row.get("Certification"), "CERTIFICATION")
        _extract(row.get("Issuing Organization"), "ORG")

        if resolved_spans := resolve_overlaps(all_spans):
            examples.append({"text": matcher.text, "entities": [s.to_spacy_format() for s in resolved_spans]})
            stats.processed_resumes += 1
            stats.entities_generated += len(resolved_spans)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        for ex in examples: f.write(json.dumps(ex) + "\n")
    
    logger.info(f"Done! Wrote {len(examples)} resumes to {OUTPUT_FILE}")

if __name__ == "__main__":
    process_dataset()
