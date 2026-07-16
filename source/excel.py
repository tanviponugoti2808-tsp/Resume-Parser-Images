import json
import glob
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
HEADER_FILL = PatternFill("solid", start_color="4472C4")
BODY_FONT = Font(name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

# Output column order. Change this list (and the matching row-building
# code in build_workbook) if you want to add/remove/reorder fields.
HEADERS = [
    "File Name",
    "Name",
    "Email",
    "Phone",
    "LinkedIn",
    "Location",
    "Overall Experience",
    "Summary",
    "Education",
    "Experience",
    "Certificates",
    "Skills",
    "Languages",
    "Achievements",
    "JSON"
]


def load_records(path_or_glob):
    """Load one or more resume JSON files. Accepts a directory, a glob
    pattern, or a single .json file. Also tolerates a file that contains
    multiple concatenated JSON objects."""
    if os.path.isdir(path_or_glob):
        files = sorted(glob.glob(os.path.join(path_or_glob, "*.json")))
    else:
        files = sorted(glob.glob(path_or_glob))

    records = []
    for f in files:
        with open(f, "r", encoding="utf-8") as fh:
            content = fh.read().strip()
        decoder = json.JSONDecoder()
        idx = 0
        while idx < len(content):
            sub = content[idx:].lstrip()
            if not sub:
                break
            idx += len(content[idx:]) - len(sub)
            obj, end = decoder.raw_decode(content, idx)
            records.append(obj)
            idx = end
    return records


def _join(values, sep="; "):
    if not values:
        return ""
    return sep.join(str(v) for v in values if v not in (None, ""))


def _flatten_education(education_list):
    """Turns the education list-of-dicts into readable multi-line text,
    one line per entry:
        Degree - Institution (Start Year - End Year)

    NOTE: keys match what parser/education_parser.py actually returns
    (Title Case with spaces), NOT schema.py's lowercase snake_case --
    those are two different naming conventions in this codebase.
    """
    if not education_list:
        return ""
    lines = []
    for edu in education_list:
        degree = str(edu.get("Degree") or "").strip()
        institution = str(edu.get("Institution") or "").strip()
        start_year = str(edu.get("Start Year") or "").strip()
        end_year = str(edu.get("End Year") or "").strip()

        parts = [p for p in [degree, institution] if p]
        line = " - ".join(parts)

        year_range = ""
        if start_year and end_year:
            year_range = f"{start_year} - {end_year}"
        elif end_year:
            year_range = end_year
        elif start_year:
            year_range = start_year

        if year_range:
            line = f"{line} ({year_range})" if line else year_range

        if line:
            lines.append(line)

    return "\n".join(lines)


def _flatten_experience(experience_list):
    """Turns the experience list-of-dicts into readable multi-line text,
    one line per entry:
        Designation - Company (Start Date - End Date)
    Responsibilities are intentionally left out of this summary cell to
    keep it scannable -- change this function if you want them included.

    NOTE: keys match what parser/experience_parser.py actually returns
    (Title Case with spaces), NOT schema.py's lowercase snake_case.
    """
    if not experience_list:
        return ""
    lines = []
    for exp in experience_list:
        designation = str(exp.get("Designation") or "").strip()
        company = str(exp.get("Company") or "").strip()
        start_date = str(exp.get("Start Date") or "").strip()
        end_date = str(exp.get("End Date") or "").strip()

        parts = [p for p in [designation, company] if p]
        line = " - ".join(parts)

        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date} - {end_date}"
        elif end_date:
            date_range = end_date
        elif start_date:
            date_range = start_date

        if date_range:
            line = f"{line} ({date_range})" if line else date_range

        if line:
            lines.append(line)

    return "\n".join(lines)


def _flatten_certificates(cert_list):
    """Turns the certificates list-of-dicts into readable multi-line text:
        Certification - Issuing Organization (Year)
    """
    if not cert_list:
        return ""
    lines = []
    for cert in cert_list:
        name = str(cert.get("Certification") or cert.get("certification") or "").strip()
        org = str(
            cert.get("Issuing_Organization")
            or cert.get("Issuing Organization")
            or cert.get("issuing_organization")
            or ""
        ).strip()
        year = str(cert.get("Year") or cert.get("year") or "").strip()

        parts = [p for p in [name, org] if p]
        line = " - ".join(parts)

        if year:
            line = f"{line} ({year})" if line else year

        if line:
            lines.append(line)

    return "\n".join(lines)


def _style_header(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, widths=None):
    """Auto-sizes each column based on the longest actual value in it,
    instead of guessing fixed widths. For columns that hold long
    multi-line flattened text (Education/Experience/Certificates/etc.),
    width is capped so one huge entry doesn't blow out the whole sheet
    -- wrap_text handles readability for those instead.

    `widths` (a list of fixed widths) can still be passed to override
    specific columns if you want manual control over any of them.
    """
    MIN_WIDTH = 10
    MAX_WIDTH = 60

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)

        if widths and col_idx <= len(widths) and widths[col_idx - 1] is not None:
            ws.column_dimensions[col_letter].width = widths[col_idx - 1]
            continue

        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                # For multi-line cells, measure the longest single line,
                # not the total character count (wrap_text handles height).
                lines = str(cell.value).split("\n")
                longest_line = max(len(line) for line in lines)
                max_len = max(max_len, longest_line)

        width = min(max(max_len + 3, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[col_letter].width = width


def _record_to_row(rec):
    """Builds the flat list of cell values (in HEADERS order) for one
    candidate record, plus the flattened text used for row-height calc."""
    education_text = _flatten_education(rec.get("education"))
    experience_text = _flatten_experience(rec.get("experience"))
    certificates_text = _flatten_certificates(rec.get("certificates"))

    row_values = [
        rec.get("file_name", ""),
        rec.get("name", ""),
        rec.get("email", ""),
        rec.get("phone", ""),
        rec.get("linkedin", ""),
        rec.get("location", ""),
        rec.get("overall_experience", ""),
        rec.get("summary", ""),
        education_text,
        experience_text,
        certificates_text,
        _join(rec.get("skills")),
        _join(rec.get("languages")),
        _join(rec.get("achievements")),
        json.dumps(rec, ensure_ascii=False),
    ]
    return row_values, education_text, experience_text, certificates_text


def _write_row(ws, row_idx, row_values, education_text, experience_text, certificates_text):
    """Writes/overwrites a single row's cell values + styling + height."""
    WRAP_COLUMNS = {8, 9, 10, 11, 12, 13, 14}  # Summary, Education, Experience,
                                                 # Certificates, Skills, Languages, Achievements

    for c, value in enumerate(row_values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=value)
        cell.font = BODY_FONT
        cell.alignment = WRAP if c in WRAP_COLUMNS else TOP

    line_count = max(
        education_text.count("\n") + 1 if education_text else 1,
        experience_text.count("\n") + 1 if experience_text else 1,
        certificates_text.count("\n") + 1 if certificates_text else 1,
    )
    ws.row_dimensions[row_idx].height = min(15 * line_count, 200)


def build_workbook(records, output_path):
    """Upserts candidate rows into the workbook at `output_path`, one row
    per candidate, keyed by File Name.

    - If `output_path` already exists, it is opened and only the rows for
      the candidates in `records` are touched: an existing candidate
      (matched by File Name) gets their row updated in place, and a new
      candidate is appended as a new row. Every other existing row is
      left completely untouched.
    - If `output_path` doesn't exist yet, a fresh workbook is created as
      before.

    This means running the pipeline for a single resume only ever
    changes that one resume's row -- it will not wipe out rows already
    written for other candidates in a previous run.
    """
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb["Candidates"] if "Candidates" in wb.sheetnames else wb.active
        ws.title = "Candidates"

        # Make sure the header row matches HEADERS (covers the case of an
        # empty/blank sheet, or a sheet that predates a header change).
        existing_header = [ws.cell(row=1, column=c).value for c in range(1, len(HEADERS) + 1)]
        if existing_header != HEADERS:
            _style_header(ws, HEADERS)

        # Build a lookup of File Name -> row index for existing rows so we
        # can update in place instead of appending duplicates.
        file_name_to_row = {}
        for row_idx in range(2, ws.max_row + 1):
            existing_file_name = ws.cell(row=row_idx, column=1).value
            if existing_file_name:
                file_name_to_row[existing_file_name] = row_idx
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        _style_header(ws, HEADERS)
        file_name_to_row = {}

    for rec in records:
        row_values, education_text, experience_text, certificates_text = _record_to_row(rec)
        file_name = row_values[0]

        if file_name and file_name in file_name_to_row:
            # Update this candidate's existing row in place; every other
            # row is left untouched.
            row_idx = file_name_to_row[file_name]
        else:
            # New candidate -- append after the last used row.
            row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
            if file_name:
                file_name_to_row[file_name] = row_idx

        _write_row(ws, row_idx, row_values, education_text, experience_text, certificates_text)

    widths = [None, None, None, None, 22, 18, 14, 35, 32, 40, 32, 30, 18, 30, 50]
    _autofit(ws, widths)

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(output_path)
    return output_path